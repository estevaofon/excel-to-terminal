#pip install openpyxl tabulate
import openpyxl
from tabulate import tabulate
path = "produtos.xlsx"

# Cria o objeto planilha
wb_obj = openpyxl.load_workbook(path)
# Seleciona a primeira aba
sheet_obj = wb_obj.active

# Numero máximo de linhas e colunas
row = sheet_obj.max_row
column = sheet_obj.max_column

# obter todas as colunas
columns = []
for column in sheet_obj.iter_cols():
    column_values = []
    for cell in column:
        column_values.append(cell.value)
    columns.append(column_values)

def print_columns(columns_list, headers=None, tablefmt=None):
    "Imprime as colunas em formato de tabela"
    table_data = list(zip(*columns_list))
    if headers:
        print(tabulate(table_data, headers=headers, tablefmt=tablefmt))
    else:
        print(tabulate(table_data, tablefmt=tablefmt))

print_columns(columns, headers='firstrow', tablefmt='fancy_grid')